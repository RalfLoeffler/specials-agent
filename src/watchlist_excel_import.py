"""Import a watchlist Excel sheet and write watchlist.yaml.

Usage:
    python -m src.watchlist_excel_import --excel watchlist.xlsx \
        --yaml watchlist.yaml

Expected columns (case-insensitive):
- name
- match_keywords (comma-separated)
- exclude_keywords (optional; comma-separated)
- stores (optional; comma-separated; blank means both stores; `none` pauses item)
- email_indices (optional; comma-separated zero-based recipient indices)
- include_unknown_half_price (optional; TRUE/FALSE/Yes/No/1/0)
- only_half_price (optional; TRUE/FALSE/Yes/No/1/0)

Requires: openpyxl, pyyaml
"""

from __future__ import annotations

import argparse
import csv
import os
from io import StringIO
from typing import Dict, List, Optional

import yaml
from openpyxl import load_workbook


def _bool_from_cell(value: object) -> bool:
    """Parse a boolean-like cell value."""
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y"}


def _cell_value(
    row: tuple[object, ...],
    header_map: Dict[str, int],
    key: str,
) -> object:
    """Return a row value for an optional column, or None when absent."""
    idx = header_map.get(key)
    return row[idx] if idx is not None and idx < len(row) else None


def _split_keywords(cell_value: object) -> List[str]:
    """Split a CSV-style keywords cell into a list, trimming blanks."""
    if cell_value is None:
        return []
    if isinstance(cell_value, list):
        return [str(x).strip() for x in cell_value if str(x).strip()]
    text = str(cell_value).strip()
    if not text:
        return []

    reader = csv.reader(StringIO(text), skipinitialspace=True)
    parts = next(reader, [])
    return [part.strip() for part in parts if part and part.strip()]


def _split_email_indices(cell_value: object) -> List[int]:
    """Split an optional CSV-style recipient index cell into integers."""
    values = _split_keywords(cell_value)
    indices: List[int] = []
    for value in values:
        try:
            index = int(value)
        except ValueError as exc:
            raise ValueError(
                f"email_indices values must be integers; got {value!r}"
            ) from exc
        if index < 0:
            raise ValueError(
                f"email_indices values must be zero or greater; got {index}"
            )
        if index not in indices:
            indices.append(index)
    return indices


def _load_existing_yaml(yaml_path: str) -> Dict[str, object]:
    """Load the existing YAML mapping so non-item keys survive import."""
    if not os.path.exists(yaml_path):
        return {}

    with open(yaml_path, "r", encoding="utf-8") as f:
        existing = yaml.safe_load(f) or {}

    if not isinstance(existing, dict):
        raise ValueError(f"{yaml_path} must contain a top-level mapping")

    return existing


def import_watchlist_from_excel(
    excel_path: str = "watchlist.xlsx",
    yaml_path: str = "watchlist.yaml",
    sheet_name: Optional[str] = None,
) -> None:
    """Read Excel and write watchlist YAML."""
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map: Dict[str, int] = {}
    for idx, name in enumerate(header_cells):
        if not name:
            continue
        key = str(name).strip().lower()
        header_map[key] = idx

    required = {
        "name",
        "match_keywords",
    }
    missing = required - set(header_map.keys())
    if missing:
        raise ValueError(f"Missing required columns: {sorted(missing)}")

    items: List[Dict[str, object]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        name = row[header_map["name"]]
        if not name:
            continue

        keywords_cell = row[header_map["match_keywords"]]
        exclude_keywords_cell = _cell_value(row, header_map, "exclude_keywords")
        stores_cell = _cell_value(row, header_map, "stores")
        email_indices_cell = _cell_value(row, header_map, "email_indices")
        include_unknown_half_price_cell = _cell_value(
            row,
            header_map,
            "include_unknown_half_price",
        )
        only_half_cell = _cell_value(row, header_map, "only_half_price")

        item: Dict[str, object] = {
            "name": str(name).strip(),
            "match_keywords": _split_keywords(keywords_cell),
            "exclude_keywords": _split_keywords(exclude_keywords_cell),
            "stores": _split_keywords(stores_cell),
            "include_unknown_half_price": (
                True
                if include_unknown_half_price_cell is None
                else _bool_from_cell(include_unknown_half_price_cell)
            ),
            "only_half_price": (
                False
                if only_half_cell is None
                else _bool_from_cell(only_half_cell)
            ),
        }
        email_indices = _split_email_indices(email_indices_cell)
        if email_indices:
            item["email_indices"] = email_indices

        items.append(item)

    data = _load_existing_yaml(yaml_path)
    data["items"] = items
    os.makedirs(os.path.dirname(yaml_path) or ".", exist_ok=True)
    with open(yaml_path, "w", encoding="utf-8") as f:
        yaml.safe_dump(data, f, sort_keys=False, allow_unicode=False)


def main():
    parser = argparse.ArgumentParser(
        description="Import an Excel watchlist and write watchlist.yaml.",
        epilog=(
            "Examples:\n"
            "  python -m src.watchlist_excel_import "
            "--excel watchlist.xlsx --yaml watchlist.yaml\n"
            "  python -m src.watchlist_excel_import "
            "--excel watchlist.xlsx --yaml watchlist.yaml "
            "--sheet watchlist"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--excel",
        default="watchlist.xlsx",
        help="Path to Excel file to read (default: watchlist.xlsx).",
    )
    parser.add_argument(
        "--yaml",
        default="watchlist.yaml",
        help="Path to YAML file to write (default: watchlist.yaml).",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Worksheet name to read (defaults to the active sheet).",
    )
    args = parser.parse_args()

    try:
        import_watchlist_from_excel(
            excel_path=args.excel,
            yaml_path=args.yaml,
            sheet_name=args.sheet,
        )
    except FileNotFoundError as exc:
        print(f"[ERROR] {exc}")
        raise SystemExit(1)

    print(
        "[INFO] Imported "
        f"{args.excel} -> {args.yaml} "
        f"({args.sheet or 'active sheet'})"
    )


if __name__ == "__main__":
    main()
